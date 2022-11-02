import io
import os
import uuid

import openpyxl
from fastapi import UploadFile, APIRouter, HTTPException
from starlette import status
from starlette.background import BackgroundTask
from starlette.responses import FileResponse

from app.main import identity_verification
from app.schemas import IdentityVerification

application = APIRouter()


@application.post("/uploadFile/")
async def create_upload_file(file: UploadFile):
    if file.filename.endswith('.xlsx'):
        f = await file.read()
        xlsx = io.BytesIO(f)
        wb = openpyxl.load_workbook(xlsx)
        ws = wb.active
        rows = ws.iter_rows()
        data = []
        req_data = []
        for cells in rows:
            row = []
            for cell in cells:
                row.append(cell.value)
            data.append(row)
        for item in data:
            if not data.index(item):
                ws['I1'].value = "IdentityVerification"
                continue
            if not item[5]:
                mn = True
            else:
                mn = False
            date = str(item[7]).strip().split("/")
            res_status = identity_verification(IdentityVerification(
                idType=str(item[1]).strip(),
                stateOfIssue=str(item[2]).strip(),
                licenseNumber=str(item[3]).strip(),
                firstName=str(item[4]).strip(),
                middleName=str(item[5]).strip(),
                familyName=str(item[6]).strip(),
                noMiddleName=mn,
                dobDay=str(date[0]).strip(),
                dobMonth=str(date[1]).strip(),
                dobYear=str(date[2]).strip()
            ))
            ws[f'I{data.index(item) + 1}'].value = res_status
        filename = f'{uuid.uuid4()}.xlsx'
        wb.save(filename)
        wb.close()
        return FileResponse(
            filename,
            filename=filename,
            background=BackgroundTask(lambda: os.remove(filename)),
        )
    else:
        raise HTTPException(status_code=status.HTTP_405_METHOD_NOT_ALLOWED)

